from flask import Flask, render_template, request, send_file, flash, redirect, url_for, session, jsonify
import os
import uuid
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import timedelta
import json

# Import our custom modules
from normalizer import normalize_and_separate_by_type
from helpers import allowed_file, clean_old_files, ensure_directories


def save_excel_clean_format(df, filepath):
    """
    Сохраняет DataFrame в Excel файл с обычным форматированием заголовков
    """
    # Создаем новую рабочую книгу
    wb = Workbook()
    ws = wb.active

    # Добавляем данные из DataFrame
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)

    # Создаем обычный стиль для всех ячеек
    normal_font = Font(
        bold=False,  # Убираем жирный
        underline='none',  # Убираем подчеркивание
        name='Calibri',  # Стандартный шрифт
        size=11,  # Стандартный размер
        color='000000'  # Черный цвет
    )

    # Убираем границы
    no_border = Border(
        left=Side(style=None),
        right=Side(style=None),
        top=Side(style=None),
        bottom=Side(style=None)
    )

    # Убираем заливку
    no_fill = PatternFill(fill_type=None)

    # Применяем стили ко всем ячейкам включая заголовки
    for row in ws.iter_rows():
        for cell in row:
            cell.font = normal_font
            cell.border = no_border
            cell.fill = no_fill

    # Сохраняем файл
    wb.save(filepath)

app = Flask(__name__)
app.secret_key = "phone_normalizer_secret_key"
app.config['SESSION_TYPE'] = 'filesystem'
app.config['PERMANENT_SESSION_LIFETIME'] = timedelta(minutes=30)

# Folders for temporary files
UPLOAD_FOLDER = 'uploads'
RESULT_FOLDER = 'results'

# Create necessary directories
ensure_directories([UPLOAD_FOLDER, RESULT_FOLDER])

app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['RESULT_FOLDER'] = RESULT_FOLDER


@app.route('/', methods=['GET', 'POST'])
def index():
    # Clean old files on main page visit
    clean_old_files(app.config)

    if request.method == 'POST':
        # Check if file was uploaded
        if 'file' not in request.files:
            flash('File not selected')
            return redirect(request.url)

        file = request.files['file']

        if file.filename == '':
            flash('File not selected')
            return redirect(request.url)

        if file and allowed_file(file.filename):
            # Generate unique filename
            from werkzeug.utils import secure_filename
            filename = secure_filename(file.filename)
            unique_filename = f"{str(uuid.uuid4())}_{filename}"
            filepath = os.path.join(app.config['UPLOAD_FOLDER'], unique_filename)

            # Save file
            file.save(filepath)

            # Get column identifiers
            phone_column = request.form.get('phone_column', '').strip()
            type_column = request.form.get('type_column', '').strip()
            sort_column = request.form.get('sort_column', '').strip() or None

            if not phone_column:
                flash('Please specify column name or number containing phone numbers')
                os.remove(filepath)  # Remove uploaded file
                return redirect(request.url)

            if not type_column:
                flash('Please specify column name or number containing participant type')
                os.remove(filepath)  # Remove uploaded file
                return redirect(request.url)

            try:
                # Process the file
                type_dfs, valid_df, invalid_df = normalize_and_separate_by_type(
                    filepath, phone_column, type_column, sort_column
                )

                # Create unique identifier for results
                from datetime import datetime
                import re

                result_id = str(uuid.uuid4())
                result_dir = os.path.join(app.config['RESULT_FOLDER'], result_id)
                os.makedirs(result_dir, exist_ok=True)

                # Define names for output files
                all_valid_filename = f"all_processed_{os.path.splitext(filename)[0]}.xlsx"
                invalid_filename = f"errors_{os.path.splitext(filename)[0]}.xlsx"

                # Save the file with all valid records
                valid_path = os.path.join(result_dir, all_valid_filename)
                save_excel_clean_format(valid_df, valid_path)

                # Save error data only if available
                has_errors = not invalid_df.empty
                if has_errors:
                    invalid_path = os.path.join(result_dir, invalid_filename)
                    save_excel_clean_format(invalid_df, invalid_path)

                # Save separate files for each participant type
                type_files = {}
                for type_name, type_df in type_dfs.items():
                    # Create a safe filename
                    safe_type_name = re.sub(r'[^\w\-_]', '_', str(type_name))
                    type_filename = f"type_{safe_type_name}_{os.path.splitext(filename)[0]}.xlsx"
                    type_path = os.path.join(result_dir, type_filename)

                    # Save the DataFrame
                    save_excel_clean_format(type_df, type_path)

                    # Save file information
                    type_files[type_name] = {
                        'filename': type_filename,
                        'count': len(type_df)
                    }

                # Save creation time information
                with open(os.path.join(result_dir, "timestamp.txt"), "w") as f:
                    f.write(datetime.now().isoformat())

                # Remove temporary file
                os.remove(filepath)

                # Save results information in session for future access
                session['result_id'] = result_id
                session['all_valid_filename'] = all_valid_filename
                session['invalid_filename'] = invalid_filename if has_errors else None
                session['original_filename'] = filename
                session['has_errors'] = has_errors
                session['error_count'] = len(invalid_df) if has_errors else 0
                session['success_count'] = len(valid_df)
                session['type_files'] = type_files

                # Redirect to results page
                return redirect(url_for('result'))

            except Exception as e:
                flash(f'File processing error: {str(e)}')
                # Remove temporary file in case of an error
                if os.path.exists(filepath):
                    os.remove(filepath)
                return redirect(request.url)
        else:
            flash('Only Excel files are allowed (.xlsx, .xls)')
            return redirect(request.url)

    return render_template('index.html')


@app.route('/result', methods=['GET'])
def result():
    """Display processing results and download links"""
    # Check if there is result data in the session
    result_id = session.get('result_id')
    if not result_id:
        flash('No processing results found. Please upload a file again.')
        return redirect(url_for('index'))

    all_valid_filename = session.get('all_valid_filename')
    invalid_filename = session.get('invalid_filename')
    original_filename = session.get('original_filename')
    has_errors = session.get('has_errors', False)
    error_count = session.get('error_count', 0)
    success_count = session.get('success_count', 0)
    type_files = session.get('type_files', {})

    # Check if result files exist
    result_dir = os.path.join(app.config['RESULT_FOLDER'], result_id)
    valid_exists = os.path.exists(os.path.join(result_dir, all_valid_filename))
    invalid_exists = invalid_filename and os.path.exists(os.path.join(result_dir, invalid_filename))

    if not valid_exists:
        flash('Result files not found. Please upload a file again.')
        return redirect(url_for('index'))

    # Check if files exist for each type
    type_files_valid = {}
    for type_name, file_info in type_files.items():
        filename = file_info['filename']
        if os.path.exists(os.path.join(result_dir, filename)):
            type_files_valid[type_name] = file_info

    return render_template(
        'result.html',
        result_id=result_id,
        all_valid_filename=all_valid_filename,
        invalid_filename=invalid_filename,
        original_filename=original_filename,
        has_errors=has_errors,
        error_count=error_count,
        success_count=success_count,
        total_count=error_count + success_count,
        type_files=type_files_valid
    )


@app.route('/download/<result_id>/<filename>', methods=['GET'])
def download_file(result_id, filename):
    """Download processed file"""
    # Check path safety
    result_dir = os.path.join(app.config['RESULT_FOLDER'], result_id)
    filepath = os.path.join(result_dir, filename)

    if not os.path.exists(filepath):
        flash('Requested file not found.')
        return redirect(url_for('index'))

    # Send file to user
    return send_file(
        filepath,
        as_attachment=True,
        download_name=filename,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )


@app.route('/preview', methods=['POST'])
def preview_columns():
    """Preview columns in uploaded file"""
    if 'file' not in request.files:
        return jsonify({'error': 'File not found'}), 400

    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': 'Empty filename'}), 400

    if not allowed_file(file.filename):
        return jsonify({'error': 'Unsupported file format'}), 400

    try:
        # Save temporary file
        from werkzeug.utils import secure_filename
        import pandas as pd

        filename = secure_filename(file.filename)
        unique_filename = f"{str(uuid.uuid4())}_{filename}"
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], unique_filename)
        file.save(filepath)

        try:
            # Read only file headers using openpyxl
            df = pd.read_excel(filepath, engine='openpyxl', nrows=0)

            # Get column list with their indices
            columns = [{'index': i, 'name': col} for i, col in enumerate(df.columns)]

            # Remove temporary file
            os.remove(filepath)

            return jsonify({'columns': columns})

        except Exception as e:
            # Remove temporary file
            if os.path.exists(filepath):
                os.remove(filepath)
            return jsonify({'error': str(e)}), 500

    except Exception as e:
        # Remove temporary file in case of an error
        if os.path.exists(filepath):
            os.remove(filepath)
        return jsonify({'error': str(e)}), 500


if __name__ == '__main__':
    app.run(debug=True)